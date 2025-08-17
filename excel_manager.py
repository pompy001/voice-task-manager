"""
Excel task manager module for Voice-Activated Task Manager
"""

import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
from datetime import datetime, date
import json

from config import EXCEL_CONFIG

logger = logging.getLogger(__name__)

class ExcelTaskManager:
    """Manages task storage and retrieval in Excel files"""
    
    def __init__(self, file_path: str = None, sheet_name: str = None):
        """
        Initialize Excel task manager
        
        Args:
            file_path: Path to Excel file
            sheet_name: Name of worksheet to use
        """
        self.file_path = file_path or EXCEL_CONFIG['file_path']
        self.sheet_name = sheet_name or EXCEL_CONFIG['sheet_name']
        self.columns = EXCEL_CONFIG['columns']
        self.priority_levels = EXCEL_CONFIG['priority_levels']
        self.status_levels = EXCEL_CONFIG['status_levels']
        
        # Ensure file exists and is properly formatted
        self._ensure_file_exists()
        self._format_worksheet()
    
    def _ensure_file_exists(self):
        """Ensure Excel file exists and create if necessary"""
        try:
            if not Path(self.file_path).exists():
                logger.info(f"Creating new Excel file: {self.file_path}")
                self._create_new_workbook()
            else:
                logger.info(f"Using existing Excel file: {self.file_path}")
                
        except Exception as e:
            logger.error(f"Error ensuring file exists: {e}")
            raise
    
    def _create_new_workbook(self):
        """Create a new Excel workbook with proper formatting"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = self.sheet_name
            
            # Add headers
            for col, header in enumerate(self.columns, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Set column widths
            column_widths = {
                'task': 40,
                'assigned_by': 15,
                'priority': 12,
                'expected_date': 15,
                'status': 12,
                'completed_date': 15,
                'created_date': 15,
                'notes': 30
            }
            
            for col, header in enumerate(self.columns, 1):
                width = column_widths.get(header, 15)
                worksheet.column_dimensions[get_column_letter(col)].width = width
            
            # Save workbook
            workbook.save(self.file_path)
            workbook.close()
            logger.info("New Excel workbook created successfully")
            
        except Exception as e:
            logger.error(f"Error creating new workbook: {e}")
            raise
    
    def _format_worksheet(self):
        """Format the worksheet with proper styling"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            # Apply conditional formatting for priorities
            self._apply_priority_formatting(worksheet)
            
            # Apply conditional formatting for status
            self._apply_status_formatting(worksheet)
            
            # Save formatting
            workbook.save(self.file_path)
            workbook.close()
            
        except Exception as e:
            logger.error(f"Error formatting worksheet: {e}")
    
    def _apply_priority_formatting(self, worksheet):
        """Apply color coding for priority levels"""
        priority_colors = {
            'urgent': 'FF0000',      # Red
            'high': 'FF6600',        # Orange
            'medium': 'FFCC00',      # Yellow
            'low': '00CC00'          # Green
        }
        
        # Find priority column
        priority_col = None
        for col, header in enumerate(self.columns, 1):
            if header == 'priority':
                priority_col = col
                break
        
        if priority_col:
            for row in range(2, worksheet.max_row + 1):
                priority_cell = worksheet.cell(row=row, column=priority_col)
                priority_value = priority_cell.value
                
                if priority_value in priority_colors:
                    color = priority_colors[priority_value]
                    priority_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    priority_cell.font = Font(bold=True, color="FFFFFF")
    
    def _apply_status_formatting(self, worksheet):
        """Apply color coding for status levels"""
        status_colors = {
            'ongoing': '0066CC',     # Blue
            'done': '00CC00',        # Green
            'paused': 'FFCC00',      # Yellow
            'cancelled': 'CC0000'    # Red
        }
        
        # Find status column
        status_col = None
        for col, header in enumerate(self.columns, 1):
            if header == 'status':
                status_col = col
                break
        
        if status_col:
            for row in range(2, worksheet.max_row + 1):
                status_cell = worksheet.cell(row=row, column=status_col)
                status_value = status_cell.value
                
                if status_value in status_colors:
                    color = status_colors[status_value]
                    status_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    status_cell.font = Font(bold=True, color="FFFFFF")
    
    def add_task(self, task_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Add a new task to the Excel file
        
        Args:
            task_data: Task data dictionary
            
        Returns:
            Result of the operation
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            # Find next empty row
            next_row = worksheet.max_row + 1
            
            # Prepare task data
            task_row = self._prepare_task_row(task_data)
            
            # Add task to worksheet
            for col, value in enumerate(task_row, 1):
                cell = worksheet.cell(row=next_row, column=col, value=value)
                
                # Apply special formatting for priority and status
                if self.columns[col-1] == 'priority' and value in self.priority_levels:
                    self._format_priority_cell(cell, value)
                elif self.columns[col-1] == 'status' and value in self.status_levels:
                    self._format_status_cell(cell, value)
            
            # Save workbook
            workbook.save(self.file_path)
            workbook.close()
            
            logger.info(f"Task added successfully at row {next_row}")
            
            return {
                'success': True,
                'row': next_row,
                'task_id': f"TASK_{next_row:04d}",
                'message': 'Task added successfully'
            }
            
        except Exception as e:
            logger.error(f"Error adding task: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': 'Failed to add task'
            }
    
    def _prepare_task_row(self, task_data: Dict[str, Any]) -> List[Any]:
        """Prepare task data for Excel row insertion"""
        row_data = []
        
        for column in self.columns:
            if column == 'created_date':
                row_data.append(datetime.now().strftime('%Y-%m-%d'))
            elif column == 'status':
                row_data.append('ongoing')  # Default status
            elif column == 'completed_date':
                row_data.append('')  # Empty for new tasks
            else:
                row_data.append(task_data.get(column, ''))
        
        return row_data
    
    def _format_priority_cell(self, cell, priority_value):
        """Format priority cell with appropriate color"""
        priority_colors = {
            'urgent': 'FF0000',
            'high': 'FF6600',
            'medium': 'FFCC00',
            'low': '00CC00'
        }
        
        if priority_value in priority_colors:
            color = priority_colors[priority_value]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def _format_status_cell(self, cell, status_value):
        """Format status cell with appropriate color"""
        status_colors = {
            'ongoing': '0066CC',
            'done': '00CC00',
            'paused': 'FFCC00',
            'cancelled': 'CC0000'
        }
        
        if status_value in status_colors:
            color = status_colors[status_value]
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
    
    def get_all_tasks(self) -> List[Dict[str, Any]]:
        """Get all tasks from the Excel file"""
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            tasks = []
            
            for row in range(2, worksheet.max_row + 1):  # Skip header row
                task = {}
                for col, header in enumerate(self.columns, 1):
                    cell_value = worksheet.cell(row=row, column=col).value
                    task[header] = cell_value
                
                # Add row number as task ID
                task['row'] = row
                task['task_id'] = f"TASK_{row:04d}"
                
                tasks.append(task)
            
            workbook.close()
            logger.info(f"Retrieved {len(tasks)} tasks")
            return tasks
            
        except Exception as e:
            logger.error(f"Error retrieving tasks: {e}")
            return []
    
    def get_tasks_by_priority(self, priority: str) -> List[Dict[str, Any]]:
        """Get tasks filtered by priority"""
        all_tasks = self.get_all_tasks()
        return [task for task in all_tasks if task.get('priority') == priority]
    
    def get_tasks_by_status(self, status: str) -> List[Dict[str, Any]]:
        """Get tasks filtered by status"""
        all_tasks = self.get_all_tasks()
        return [task for task in all_tasks if task.get('status') == status]
    
    def get_next_priority_task(self) -> Optional[Dict[str, Any]]:
        """Get the next highest priority task"""
        all_tasks = self.get_all_tasks()
        
        # Filter ongoing tasks
        ongoing_tasks = [task for task in all_tasks if task.get('status') == 'ongoing']
        
        if not ongoing_tasks:
            return None
        
        # Sort by priority and expected date
        priority_order = {'urgent': 0, 'high': 1, 'medium': 2, 'low': 3}
        
        def sort_key(task):
            priority = task.get('priority', 'low')
            expected_date = task.get('expected_date', '')
            
            # Parse date if available
            try:
                if expected_date:
                    date_obj = datetime.strptime(str(expected_date), '%Y-%m-%d').date()
                    days_until_due = (date_obj - date.today()).days
                else:
                    days_until_due = 999  # No date = low priority
            except:
                days_until_due = 999
            
            return (priority_order.get(priority, 3), days_until_due)
        
        sorted_tasks = sorted(ongoing_tasks, key=sort_key)
        return sorted_tasks[0] if sorted_tasks else None
    
    def update_task_status(self, task_id: str, new_status: str) -> Dict[str, Any]:
        """
        Update task status
        
        Args:
            task_id: Task identifier (e.g., "TASK_0001")
            new_status: New status value
            
        Returns:
            Result of the operation
        """
        try:
            if new_status not in self.status_levels:
                return {
                    'success': False,
                    'error': f'Invalid status: {new_status}',
                    'message': f'Status must be one of: {", ".join(self.status_levels)}'
                }
            
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            # Parse task ID to get row number
            try:
                row_num = int(task_id.split('_')[1])
            except (IndexError, ValueError):
                return {
                    'success': False,
                    'error': f'Invalid task ID format: {task_id}',
                    'message': 'Task ID must be in format TASK_XXXX'
                }
            
            # Find status column
            status_col = None
            for col, header in enumerate(self.columns, 1):
                if header == 'status':
                    status_col = col
                    break
            
            if not status_col:
                return {
                    'success': False,
                    'error': 'Status column not found',
                    'message': 'Worksheet format error'
                }
            
            # Update status
            status_cell = worksheet.cell(row=row_num, column=status_col)
            status_cell.value = new_status
            
            # Apply formatting
            self._format_status_cell(status_cell, new_status)
            
            # Update completed date if status is 'done'
            if new_status == 'done':
                completed_date_col = None
                for col, header in enumerate(self.columns, 1):
                    if header == 'completed_date':
                        completed_date_col = col
                        break
                
                if completed_date_col:
                    completed_cell = worksheet.cell(row=row_num, column=completed_date_col)
                    completed_cell.value = datetime.now().strftime('%Y-%m-%d')
            
            # Save workbook
            workbook.save(self.file_path)
            workbook.close()
            
            logger.info(f"Task {task_id} status updated to {new_status}")
            
            return {
                'success': True,
                'task_id': task_id,
                'new_status': new_status,
                'message': f'Task status updated to {new_status}'
            }
            
        except Exception as e:
            logger.error(f"Error updating task status: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': 'Failed to update task status'
            }
    
    def delete_task(self, task_id: str) -> Dict[str, Any]:
        """
        Delete a task from the Excel file
        
        Args:
            task_id: Task identifier
            
        Returns:
            Result of the operation
        """
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            worksheet = workbook[self.sheet_name]
            
            # Parse task ID to get row number
            try:
                row_num = int(task_id.split('_')[1])
            except (IndexError, ValueError):
                return {
                    'success': False,
                    'error': f'Invalid task ID format: {task_id}',
                    'message': 'Task ID must be in format TASK_XXXX'
                }
            
            # Delete the row
            worksheet.delete_rows(row_num)
            
            # Save workbook
            workbook.save(self.file_path)
            workbook.close()
            
            logger.info(f"Task {task_id} deleted successfully")
            
            return {
                'success': True,
                'task_id': task_id,
                'message': 'Task deleted successfully'
            }
            
        except Exception as e:
            logger.error(f"Error deleting task: {e}")
            return {
                'success': False,
                'error': str(e),
                'message': 'Failed to delete task'
            }
    
    def get_task_statistics(self) -> Dict[str, Any]:
        """Get statistics about tasks"""
        try:
            all_tasks = self.get_all_tasks()
            
            stats = {
                'total_tasks': len(all_tasks),
                'by_priority': {},
                'by_status': {},
                'overdue_tasks': 0,
                'due_today': 0,
                'due_this_week': 0
            }
            
            today = date.today()
            
            for task in all_tasks:
                # Count by priority
                priority = task.get('priority', 'unknown')
                stats['by_priority'][priority] = stats['by_priority'].get(priority, 0) + 1
                
                # Count by status
                status = task.get('status', 'unknown')
                stats['by_status'][status] = stats['by_status'].get(status, 0) + 1
                
                # Check dates
                expected_date = task.get('expected_date')
                if expected_date and task.get('status') == 'ongoing':
                    try:
                        due_date = datetime.strptime(str(expected_date), '%Y-%m-%d').date()
                        days_until_due = (due_date - today).days
                        
                        if days_until_due < 0:
                            stats['overdue_tasks'] += 1
                        elif days_until_due == 0:
                            stats['due_today'] += 1
                        elif days_until_due <= 7:
                            stats['due_this_week'] += 1
                    except:
                        pass
            
            return stats
            
        except Exception as e:
            logger.error(f"Error getting task statistics: {e}")
            return {}
    
    def cleanup(self):
        """Clean up resources"""
        # No specific cleanup needed for openpyxl
        logger.info("Excel task manager cleaned up")


if __name__ == "__main__":
    # Test Excel manager functionality
    import logging
    logging.basicConfig(level=logging.INFO)
    
    # Create a test Excel file
    test_file = "test_tasks.xlsx"
    
    try:
        manager = ExcelTaskManager(test_file)
        
        # Test adding a task
        test_task = {
            'task': 'Test dashboard project',
            'assigned_by': 'Test User',
            'priority': 'high',
            'expected_date': '2024-07-04',
            'notes': 'This is a test task'
        }
        
        result = manager.add_task(test_task)
        print(f"Add task result: {result}")
        
        # Test getting all tasks
        tasks = manager.get_all_tasks()
        print(f"Retrieved {len(tasks)} tasks")
        
        # Test getting next priority task
        next_task = manager.get_next_priority_task()
        print(f"Next priority task: {next_task}")
        
        # Test statistics
        stats = manager.get_task_statistics()
        print(f"Task statistics: {stats}")
        
        # Clean up test file
        Path(test_file).unlink(missing_ok=True)
        
    except Exception as e:
        print(f"Error during testing: {e}")
        # Clean up test file on error
        Path(test_file).unlink(missing_ok=True)

